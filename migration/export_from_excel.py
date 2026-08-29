#!/usr/bin/env python3
"""
Migrates the legacy 'FY 26 Review Sheet' workbook into a clean seed.json
for the Municipal Client Tracker app.

Source of truth for roster + assignment: the 16 advisor tabs (1-Sue ... 16-Michelle),
since that's what the team actually works from day to day. The Territory List
sheet supplies municipal_type / county / client-vs-prospect status per community.
"""
import json
import re
import sys
import datetime
import openpyxl

SRC = "/root/.claude/uploads/b9f2a147-d425-516a-a954-3810aef889dd/915250de-FY_26_Review_Sheet_extra.xlsx"
OUT = "/home/claude/muni-tracker/migration/seed.json"

TAB_TO_REGION_ADVISOR = {
    "1-Sue": (1, "Sue"),
    "2-Sue": (2, "Sue"),
    "3-Sue": (3, "Sue"),
    "4-Sue": (4, "Sue"),
    "5-Brian": (5, "Brian"),
    "6-Sue": (6, "Sue"),
    "7-Brian": (7, "Brian"),
    "8-Kath": (8, "Kath"),
    "9-Kath": (9, "Kath"),
    "10-Brian": (10, "Brian"),
    "11-B,M,K": (11, "B,M,K"),
    "12-Kath": (12, "Kath"),
    "13-Michelle": (13, "Michelle"),
    "14-Michelle": (14, "Michelle"),
    "15-Michelle": (15, "Michelle"),
    "16-Michelle": (16, "Michelle"),
}

DATE_LIKE = re.compile(r"^\d{4}-\d{2}-\d{2}$|^\d{1,2}/\d{1,2}(/\d{2,4})?$")

def looks_bogus(s):
    """Catch column-shift artifacts: a City/Community cell that's actually a
    stray date or a lone number (seen e.g. on the Goshen row, which was
    manually moved between tabs and lost its City/Community columns)."""
    if s is None:
        return True
    s = str(s).strip()
    if not s:
        return True
    if DATE_LIKE.match(s):
        return True
    if re.fullmatch(r"\d{1,2}", s):
        return True
    return False

def clean(v):
    """Turn spreadsheet junk values into either None or a trimmed string."""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.time):
        # A bare time(0,0) in these sheets is a formatting artifact for "blank", not real data
        if v.hour == 0 and v.minute == 0:
            return None
        return v.strftime("%H:%M")
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s in ("", "-", "--", "#N/A", "N/A"):
        return None
    return s

def parse_date_guess(text):
    """Best-effort extraction of an ISO date from messy strings like '9/17 @ 11:30 (lunch)'."""
    if not text or not isinstance(text, str):
        return None
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text.strip()):
        return text.strip()
    m = re.match(r"^\s*(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?", text)
    if not m:
        return None
    month, day, year = m.groups()
    month, day = int(month), int(day)
    if year:
        year = int(year)
        if year < 100:
            year += 2000
    else:
        # FY26 sheet: Jan-Jun dates are 2026, Jul-Dec dates are 2025 (fiscal year school-district style)
        year = 2026 if month <= 6 else 2025
    try:
        return datetime.date(year, month, day).isoformat()
    except ValueError:
        return None

def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    # ---- Territory List -> community metadata ----
    communities = {}
    ws = wb["Territory List"]
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = clean(row[3])
        if not name:
            continue
        communities[name.strip().lower()] = {
            "name": name.strip(),
            "municipal_type": clean(row[4]) or "Town",
            "county": clean(row[5]),
            "region": row[6],
            "status": "client" if row[1] else ("prospect" if row[2] else "client"),
        }

    # ---- Master "DO NOT EDIT" list -> household -> (city, community) lookup ----
    # Several advisor tabs (5,7,8,9,10,11,12,13,14,15,16) don't carry their own
    # City/Community columns at all, so we backfill from the master list.
    master_lookup = {}
    ws = wb["DO NOT EDIT - Client List"]
    mheaders = [str(c.value).strip() if c.value else "" for c in ws[1]]
    midx = {h: i for i, h in enumerate(mheaders) if h}
    for row in ws.iter_rows(min_row=2, values_only=True):
        hh = clean(row[midx.get("Household Name")]) if midx.get("Household Name") is not None else None
        if not hh:
            continue
        master_lookup[hh.strip().upper()] = {
            "city": clean(row[midx.get("City")]) if midx.get("City") is not None else None,
            "community": clean(row[midx.get("Community")]) if midx.get("Community") is not None else None,
        }

    # Manual fixes for entities added ad hoc (per the workbook's own "CHANGES needed"
    # tab) that never got linked to a community/city.
    MANUAL_COMMUNITY = {
        "PORTER MEMORIAL LIBRARY": ("Blandford", "Blandford"),
        "LAKE WICKABOAG PRESERVATION": ("West Brookfield", "West Brookfield"),
        "OXFORD LIBRARY": ("Oxford", "Oxford"),
        "SUTTON OPEB COMMITTEE": ("Sutton", "Sutton"),
        "WORCESTER LIBRARY": ("Worcester", "Worcester"),
        "WORCESTER, CITY OF": ("Worcester", "Worcester"),
    }

    # ---- 16 advisor tabs -> client roster ----
    clients = []
    seen = set()
    for tab, (region, advisor) in TAB_TO_REGION_ADVISOR.items():
        ws = wb[tab]
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        idx = {h: i for i, h in enumerate(headers) if h}

        def get(row, key):
            i = idx.get(key)
            return clean(row[i]) if i is not None and i < len(row) else None

        for row in ws.iter_rows(min_row=2, values_only=True):
            household = get(row, "Household Name")
            if not household:
                continue
            household = household.strip()
            community = get(row, "Community")
            city = get(row, "City")
            if looks_bogus(community):
                community = None
            if looks_bogus(city):
                city = None
            if not community or not city:
                fallback = master_lookup.get(household.upper())
                if fallback:
                    community = community or fallback["community"]
                    city = city or fallback["city"]
            if (not community or not city) and household.upper() in MANUAL_COMMUNITY:
                m_comm, m_city = MANUAL_COMMUNITY[household.upper()]
                community = community or m_comm
                city = city or m_city
            last_review_text = get(row, "Last Review")
            special_notes = get(row, "Special Notes")
            material_raw = get(row, "Copy/Material Count")
            coverage_note = get(row, "Advisor Number/Kathleen?")
            next_review_text = get(row, "NEW review date") or get(row, "New Review Date")
            done_raw = get(row, "Done?")

            # Copy/Material Count column was frequently misused for free-text rep notes
            # (per team's own "CHANGES needed" list, item #1). Split it here.
            material_count = None
            assigned_rep_note = None
            if material_raw is not None:
                s = str(material_raw)
                if re.fullmatch(r"\d+", s):
                    material_count = int(s)
                else:
                    assigned_rep_note = s

            key = (household.upper(), region)
            if key in seen:
                continue
            seen.add(key)

            clients.append({
                "household_name": household,
                "community": community or city or household,
                "mailing_city": city,
                "region": region,
                "advisor": advisor,
                "last_review_text": last_review_text,
                "last_review_date": parse_date_guess(last_review_text) if last_review_text else None,
                "next_review_text": next_review_text,
                "next_review_date": parse_date_guess(next_review_text) if next_review_text else None,
                "special_notes": special_notes,
                "material_count": material_count,
                "assigned_rep_note": assigned_rep_note,
                "coverage_note": coverage_note,
                "done": bool(done_raw) and done_raw not in (0, "0"),
                "active": True,
            })

    # Fill in any community referenced by a client but missing from Territory List
    for c in clients:
        key = c["community"].strip().lower()
        if key not in communities:
            communities[key] = {
                "name": c["community"].strip(),
                "municipal_type": "Town",
                "county": None,
                "region": c["region"],
                "status": "client",
            }

    seed = {
        "generated_at": datetime.datetime.utcnow().isoformat() + "Z",
        "source_file": "FY 26 Review Sheet extra.xlsx",
        "regions": [{"id": r, "advisor": a} for tab, (r, a) in TAB_TO_REGION_ADVISOR.items()],
        "communities": list(communities.values()),
        "clients": clients,
    }

    with open(OUT, "w") as f:
        json.dump(seed, f, indent=2)

    print(f"Wrote {len(clients)} clients across {len(communities)} communities to {OUT}")

if __name__ == "__main__":
    main()
